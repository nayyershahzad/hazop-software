from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List, Optional
from uuid import UUID
from datetime import datetime
import uuid as uuid_module
from app.database import get_db
from app.models.study import HazopStudy, HazopNode, Deviation
from app.models.hazop_entities import Cause, Consequence, Safeguard, Recommendation
from app.models.user import User
from app.api.deps import get_current_user

router = APIRouter(prefix="/api/studies", tags=["studies"])

# Schemas
class CreateStudyRequest(BaseModel):
    title: str
    description: Optional[str] = None
    facility_name: Optional[str] = None

class StudyResponse(BaseModel):
    id: UUID
    title: str
    description: Optional[str]
    facility_name: Optional[str]
    status: str
    created_at: datetime
    node_count: Optional[int] = 0

    class Config:
        from_attributes = True

class CreateNodeRequest(BaseModel):
    node_number: str
    node_name: str
    description: Optional[str] = None
    design_intent: Optional[str] = None

class NodeResponse(BaseModel):
    id: UUID
    study_id: UUID
    node_number: str
    node_name: str
    description: Optional[str]
    design_intent: Optional[str]
    status: str
    deviation_count: Optional[int] = 0

    class Config:
        from_attributes = True

class CreateDeviationRequest(BaseModel):
    parameter: str
    guide_word: str
    deviation_description: str

class DeviationResponse(BaseModel):
    id: UUID
    node_id: UUID
    parameter: str
    guide_word: str
    deviation_description: str

    class Config:
        from_attributes = True

# Study endpoints
@router.post("/", response_model=StudyResponse)
def create_study(
    req: CreateStudyRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new HAZOP study (organization-isolated)"""
    study = HazopStudy(
        title=req.title,
        description=req.description,
        facility_name=req.facility_name,
        created_by=current_user.id,
        organization_id=current_user.organization_id  # Auto-set from current user
    )
    db.add(study)
    db.commit()
    db.refresh(study)
    return study

from app.api.pagination import PaginationParams, PaginatedResponse

@router.get("/", response_model=PaginatedResponse[StudyResponse])
def list_studies(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    pagination: PaginationParams = Depends(),
):
    """
    List HAZOP studies with pagination and field selection.

    - Supports pagination with skip/limit parameters
    - Supports field selection with the fields parameter
    - Results are ordered by most recent first
    """
    # Base query with organization isolation
    base_query = db.query(HazopStudy).filter(
        HazopStudy.organization_id == current_user.organization_id
    )

    # Get total count for pagination
    total_count = base_query.count()

    # Apply pagination
    studies = base_query.order_by(
        HazopStudy.created_at.desc()  # Most recent first
    ).offset(pagination.skip).limit(pagination.limit).all()

    # Only fetch node counts if we have studies
    study_ids = [study.id for study in studies]

    # Prepare metadata for the response
    metadata = {"execution_time_ms": 0}

    if study_ids:
        # Start time tracking for performance metrics
        import time
        start_time = time.time()

        # Get node counts per study in a single query
        from sqlalchemy.sql import func
        node_counts = dict(
            db.query(
                HazopNode.study_id,
                func.count(HazopNode.id).label("count")
            ).filter(
                HazopNode.study_id.in_(study_ids)
            ).group_by(
                HazopNode.study_id
            ).all()
        )

        # Attach the counts to study objects
        for study in studies:
            study.node_count = node_counts.get(study.id, 0)

        # Calculate execution time
        metadata["execution_time_ms"] = round((time.time() - start_time) * 1000, 2)

    # Create paginated response
    return PaginatedResponse.create(
        items=studies,
        total_count=total_count,
        params=pagination,
        metadata=metadata
    )

@router.get("/{study_id}", response_model=StudyResponse)
def get_study(
    study_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get a specific study (organization-isolated)"""
    study = db.query(HazopStudy).filter(
        HazopStudy.id == study_id,
        HazopStudy.organization_id == current_user.organization_id  # Data isolation
    ).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found or access denied")
    return study

# Node endpoints
@router.post("/{study_id}/nodes", response_model=NodeResponse)
def create_node(
    study_id: UUID,
    req: CreateNodeRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Add a node to a study"""
    study = db.query(HazopStudy).filter(HazopStudy.id == study_id).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    node = HazopNode(
        study_id=study_id,
        node_number=req.node_number,
        node_name=req.node_name,
        description=req.description,
        design_intent=req.design_intent
    )
    db.add(node)
    db.commit()
    db.refresh(node)
    return node

@router.get("/{study_id}/nodes", response_model=PaginatedResponse[NodeResponse])
def list_nodes(
    study_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    pagination: PaginationParams = Depends(),
):
    """
    List all nodes in a study with pagination and optimization

    - Supports pagination with skip/limit parameters
    - Supports field selection with fields parameter
    - Optimized query with deviation counts
    """
    # Verify study exists and belongs to user's organization
    study = db.query(HazopStudy).filter(
        HazopStudy.id == study_id,
        HazopStudy.organization_id == current_user.organization_id
    ).first()

    if not study:
        raise HTTPException(status_code=404, detail="Study not found or access denied")

    # Start time tracking for performance metrics
    import time
    start_time = time.time()

    # Query nodes with deviation count optimization
    from sqlalchemy.sql import func

    # First, get the total count for pagination
    base_query = db.query(HazopNode).filter(HazopNode.study_id == study_id)
    total_count = base_query.count()

    # Get the nodes with counts, applying pagination
    nodes_with_counts = db.query(
        HazopNode,
        func.count(Deviation.id).label("deviation_count")
    ).outerjoin(
        Deviation, Deviation.node_id == HazopNode.id
    ).filter(
        HazopNode.study_id == study_id
    ).group_by(
        HazopNode.id
    ).order_by(
        HazopNode.node_number
    ).offset(pagination.skip).limit(pagination.limit).all()

    # Attach deviation counts to node objects
    result = []
    for node, deviation_count in nodes_with_counts:
        node.deviation_count = deviation_count
        result.append(node)

    # Calculate execution time for metadata
    execution_time = round((time.time() - start_time) * 1000, 2)

    # Create paginated response
    return PaginatedResponse.create(
        items=result,
        total_count=total_count,
        params=pagination,
        metadata={"execution_time_ms": execution_time}
    )

# Deviation endpoints
@router.post("/nodes/{node_id}/deviations", response_model=DeviationResponse)
def create_deviation(
    node_id: UUID,
    req: CreateDeviationRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Add a deviation to a node"""
    node = db.query(HazopNode).filter(HazopNode.id == node_id).first()
    if not node:
        raise HTTPException(status_code=404, detail="Node not found")

    deviation = Deviation(
        node_id=node_id,
        parameter=req.parameter,
        guide_word=req.guide_word,
        deviation_description=req.deviation_description
    )
    db.add(deviation)
    db.commit()
    db.refresh(deviation)
    return deviation

@router.get("/nodes/{node_id}/deviations", response_model=List[DeviationResponse])
def list_deviations(
    node_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List all deviations for a node"""
    deviations = db.query(Deviation).filter(Deviation.node_id == node_id).all()
    return deviations

@router.delete("/{study_id}")
def delete_study(
    study_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a study and all its related data"""
    study = db.query(HazopStudy).filter(HazopStudy.id == study_id).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    db.delete(study)
    db.commit()
    return {"message": "Study deleted successfully"}

@router.put("/nodes/{node_id}", response_model=NodeResponse)
def update_node(
    node_id: UUID,
    req: CreateNodeRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update a node's information"""
    node = db.query(HazopNode).filter(HazopNode.id == node_id).first()
    if not node:
        raise HTTPException(status_code=404, detail="Node not found")

    # Update node fields
    node.node_number = req.node_number
    node.node_name = req.node_name
    node.description = req.description
    node.design_intent = req.design_intent

    db.commit()
    db.refresh(node)
    return node

@router.delete("/nodes/{node_id}")
def delete_node(
    node_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a node and all its deviations"""
    node = db.query(HazopNode).filter(HazopNode.id == node_id).first()
    if not node:
        raise HTTPException(status_code=404, detail="Node not found")

    db.delete(node)
    db.commit()
    return {"message": "Node deleted successfully"}

@router.delete("/deviations/{deviation_id}")
def delete_deviation(
    deviation_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a deviation and all its HAZOP analysis data"""
    deviation = db.query(Deviation).filter(Deviation.id == deviation_id).first()
    if not deviation:
        raise HTTPException(status_code=404, detail="Deviation not found")

    db.delete(deviation)
    db.commit()
    return {"message": "Deviation deleted successfully"}

# Duplicate/Copy endpoints
class DuplicateNodeRequest(BaseModel):
    new_node_number: str
    new_node_name: str
    include_deviations: bool = True
    include_causes: bool = True
    include_consequences: bool = True
    include_safeguards: bool = True
    include_recommendations: bool = True

@router.post("/nodes/{node_id}/duplicate", response_model=NodeResponse)
def duplicate_node(
    node_id: UUID,
    req: DuplicateNodeRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Duplicate a node with all its data"""
    # Get source node
    source_node = db.query(HazopNode).filter(HazopNode.id == node_id).first()
    if not source_node:
        raise HTTPException(status_code=404, detail="Node not found")

    # Create new node
    new_node = HazopNode(
        study_id=source_node.study_id,
        node_number=req.new_node_number,
        node_name=req.new_node_name,
        description=source_node.description,
        design_intent=source_node.design_intent,
        status="pending"
    )
    db.add(new_node)
    db.flush()

    # Copy deviations if requested
    if req.include_deviations:
        source_deviations = db.query(Deviation).filter(
            Deviation.node_id == node_id
        ).all()

        for src_dev in source_deviations:
            new_dev = Deviation(
                node_id=new_node.id,
                parameter=src_dev.parameter,
                guide_word=src_dev.guide_word,
                deviation_description=src_dev.deviation_description
            )
            db.add(new_dev)
            db.flush()

            # Copy causes
            if req.include_causes:
                src_causes = db.query(Cause).filter(
                    Cause.deviation_id == src_dev.id
                ).all()
                for src_cause in src_causes:
                    new_cause = Cause(
                        deviation_id=new_dev.id,
                        cause_description=src_cause.cause_description,
                        likelihood=src_cause.likelihood,
                        created_by=current_user.id
                    )
                    db.add(new_cause)

            # Copy consequences
            if req.include_consequences:
                src_consequences = db.query(Consequence).filter(
                    Consequence.deviation_id == src_dev.id
                ).all()
                for src_cons in src_consequences:
                    new_cons = Consequence(
                        deviation_id=new_dev.id,
                        consequence_description=src_cons.consequence_description,
                        severity=src_cons.severity,
                        category=src_cons.category,
                        created_by=current_user.id
                    )
                    db.add(new_cons)

            # Copy safeguards
            if req.include_safeguards:
                src_safeguards = db.query(Safeguard).filter(
                    Safeguard.deviation_id == src_dev.id
                ).all()
                for src_safe in src_safeguards:
                    new_safe = Safeguard(
                        deviation_id=new_dev.id,
                        safeguard_description=src_safe.safeguard_description,
                        safeguard_type=src_safe.safeguard_type,
                        effectiveness=src_safe.effectiveness,
                        created_by=current_user.id
                    )
                    db.add(new_safe)

            # Copy recommendations
            if req.include_recommendations:
                src_recs = db.query(Recommendation).filter(
                    Recommendation.deviation_id == src_dev.id
                ).all()
                for src_rec in src_recs:
                    new_rec = Recommendation(
                        deviation_id=new_dev.id,
                        recommendation_description=src_rec.recommendation_description,
                        priority=src_rec.priority,
                        responsible_party=src_rec.responsible_party,
                        target_date=src_rec.target_date,
                        status="open",
                        created_by=current_user.id
                    )
                    db.add(new_rec)

    db.commit()
    db.refresh(new_node)
    return new_node

class SimilarDeviationResponse(BaseModel):
    deviation_id: UUID
    node_id: UUID
    node_number: str
    node_name: str
    study_id: UUID
    study_name: str
    parameter: str
    guide_word: str
    deviation_description: str
    causes_count: int
    consequences_count: int
    safeguards_count: int
    recommendations_count: int

    class Config:
        from_attributes = True

@router.get("/deviations/similar", response_model=List[SimilarDeviationResponse])
def find_similar_deviations(
    parameter: str,
    guide_word: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Find similar deviations from other nodes/studies"""
    # Find deviations with same parameter and guide word
    deviations = db.query(
        Deviation,
        HazopNode,
        HazopStudy
    ).join(
        HazopNode, Deviation.node_id == HazopNode.id
    ).join(
        HazopStudy, HazopNode.study_id == HazopStudy.id
    ).filter(
        Deviation.parameter == parameter,
        Deviation.guide_word == guide_word
    ).all()

    results = []
    for dev, node, study in deviations:
        # Count related items
        causes_count = db.query(Cause).filter(Cause.deviation_id == dev.id).count()
        consequences_count = db.query(Consequence).filter(Consequence.deviation_id == dev.id).count()
        safeguards_count = db.query(Safeguard).filter(Safeguard.deviation_id == dev.id).count()
        recommendations_count = db.query(Recommendation).filter(Recommendation.deviation_id == dev.id).count()

        # Only include if it has some data
        if causes_count > 0 or consequences_count > 0 or safeguards_count > 0:
            results.append({
                "deviation_id": dev.id,
                "node_id": node.id,
                "node_number": node.node_number,
                "node_name": node.node_name,
                "study_id": study.id,
                "study_name": study.title,
                "parameter": dev.parameter,
                "guide_word": dev.guide_word,
                "deviation_description": dev.deviation_description,
                "causes_count": causes_count,
                "consequences_count": consequences_count,
                "safeguards_count": safeguards_count,
                "recommendations_count": recommendations_count
            })

    return results

class CopyFromPreviousRequest(BaseModel):
    source_deviation_id: UUID
    copy_causes: bool = True
    copy_consequences: bool = True
    copy_safeguards: bool = True
    copy_recommendations: bool = False

@router.post("/deviations/{deviation_id}/copy-from-previous")
def copy_from_previous_deviation(
    deviation_id: UUID,
    req: CopyFromPreviousRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Copy data from a previous similar deviation"""
    # Get target deviation
    target_deviation = db.query(Deviation).filter(
        Deviation.id == deviation_id
    ).first()
    if not target_deviation:
        raise HTTPException(status_code=404, detail="Target deviation not found")

    # Get source deviation
    source_deviation = db.query(Deviation).filter(
        Deviation.id == req.source_deviation_id
    ).first()
    if not source_deviation:
        raise HTTPException(status_code=404, detail="Source deviation not found")

    copied_counts = {
        "causes": 0,
        "consequences": 0,
        "safeguards": 0,
        "recommendations": 0
    }

    # Copy causes
    if req.copy_causes:
        src_causes = db.query(Cause).filter(
            Cause.deviation_id == source_deviation.id
        ).all()
        for src_cause in src_causes:
            new_cause = Cause(
                deviation_id=target_deviation.id,
                cause_description=src_cause.cause_description,
                likelihood=src_cause.likelihood,
                created_by=current_user.id
            )
            db.add(new_cause)
            copied_counts["causes"] += 1

    # Copy consequences
    if req.copy_consequences:
        src_consequences = db.query(Consequence).filter(
            Consequence.deviation_id == source_deviation.id
        ).all()
        for src_cons in src_consequences:
            new_cons = Consequence(
                deviation_id=target_deviation.id,
                consequence_description=src_cons.consequence_description,
                severity=src_cons.severity,
                category=src_cons.category,
                created_by=current_user.id
            )
            db.add(new_cons)
            copied_counts["consequences"] += 1

    # Copy safeguards
    if req.copy_safeguards:
        src_safeguards = db.query(Safeguard).filter(
            Safeguard.deviation_id == source_deviation.id
        ).all()
        for src_safe in src_safeguards:
            new_safe = Safeguard(
                deviation_id=target_deviation.id,
                safeguard_description=src_safe.safeguard_description,
                safeguard_type=src_safe.safeguard_type,
                effectiveness=src_safe.effectiveness,
                created_by=current_user.id
            )
            db.add(new_safe)
            copied_counts["safeguards"] += 1

    # Copy recommendations
    if req.copy_recommendations:
        src_recs = db.query(Recommendation).filter(
            Recommendation.deviation_id == source_deviation.id
        ).all()
        for src_rec in src_recs:
            new_rec = Recommendation(
                deviation_id=target_deviation.id,
                recommendation_description=src_rec.recommendation_description,
                priority=src_rec.priority,
                responsible_party=src_rec.responsible_party,
                target_date=src_rec.target_date,
                status="open",
                created_by=current_user.id
            )
            db.add(new_rec)
            copied_counts["recommendations"] += 1

    db.commit()

    return {
        "message": "Data copied successfully",
        "copied": copied_counts
    }

# Dashboard endpoint
from app.models.risk_assessment import ImpactAssessment
from sqlalchemy.sql import func
from sqlalchemy import distinct, case

@router.get("/{study_id}/dashboard")
def get_study_dashboard(
    study_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get dashboard data for a study"""
    # Cache in variables to avoid multiple lookups
    import time
    start_time = time.time()

    study = db.query(HazopStudy).filter(HazopStudy.id == study_id).first()
    if not study:
        raise HTTPException(status_code=404, detail="Study not found")

    # Use a single query with JOIN to get counts by node
    node_deviation_query = db.query(
        HazopNode.id.label('node_id'),
        HazopNode.node_name.label('node_name'),
        func.count(distinct(Deviation.id)).label('deviation_count')
    ).join(
        Deviation, Deviation.node_id == HazopNode.id, isouter=True
    ).filter(
        HazopNode.study_id == study_id
    ).group_by(
        HazopNode.id, HazopNode.node_name
    )

    # Execute query and format results
    nodes_with_counts = node_deviation_query.all()
    deviations_by_node = [
        {
            "node_id": str(row.node_id),
            "node_name": row.node_name,
            "count": row.deviation_count
        }
        for row in nodes_with_counts
    ]

    # Sort by count descending
    deviations_by_node.sort(key=lambda x: x["count"], reverse=True)

    # Get total counts efficiently with subqueries
    # Count nodes directly
    total_nodes = db.query(func.count(HazopNode.id)).filter(
        HazopNode.study_id == study_id
    ).scalar()

    # Create efficient subquery for nodes in this study
    nodes_subq = db.query(HazopNode.id).filter(
        HazopNode.study_id == study_id
    ).subquery()

    # Count entities using subquery - more efficient than multiple joins
    total_deviations = db.query(func.count(Deviation.id)).filter(
        Deviation.node_id.in_(nodes_subq)
    ).scalar()

    # Create deviations subquery
    deviations_subq = db.query(Deviation.id).filter(
        Deviation.node_id.in_(nodes_subq)
    ).subquery()

    # Count entities using subqueries
    total_causes = db.query(func.count(Cause.id)).filter(
        Cause.deviation_id.in_(deviations_subq)
    ).scalar()

    # Create causes subquery
    causes_subq = db.query(Cause.id).filter(
        Cause.deviation_id.in_(deviations_subq)
    ).subquery()

    # Count consequences using subquery
    total_consequences = db.query(func.count(Consequence.id)).filter(
        Consequence.deviation_id.in_(deviations_subq)
    ).scalar()

    # Create consequences subquery
    consequences_subq = db.query(Consequence.id).filter(
        Consequence.deviation_id.in_(deviations_subq)
    ).subquery()

    # Count safeguards using subquery
    total_safeguards = db.query(func.count(Safeguard.id)).filter(
        Safeguard.consequence_id.in_(consequences_subq)
    ).scalar()

    # Count recommendations
    total_recommendations = db.query(func.count(Recommendation.id)).filter(
        Recommendation.consequence_id.in_(consequences_subq)
    ).scalar()

    # Get risk distribution with a single query using CASE statements
    risk_query = db.query(
        ImpactAssessment.risk_level.label('level'),
        func.count(ImpactAssessment.id).label('count')
    ).filter(
        ImpactAssessment.consequence_id.in_(consequences_subq)
    ).group_by(
        ImpactAssessment.risk_level
    )

    # Initialize risk distribution
    risk_distribution = {"critical": 0, "high": 0, "medium": 0, "low": 0}

    # Populate with actual data
    for row in risk_query:
        if row.level and row.level.lower() in risk_distribution:
            risk_distribution[row.level.lower()] = row.count

    end_time = time.time()
    execution_time = end_time - start_time

    return {
        "study": {
            "id": str(study.id),
            "title": study.title,
            "facility_name": study.facility_name
        },
        "metrics": {
            "total_nodes": total_nodes,
            "total_deviations": total_deviations,
            "total_causes": total_causes,
            "total_consequences": total_consequences,
            "total_safeguards": total_safeguards,
            "total_recommendations": total_recommendations,
            "risk_distribution": risk_distribution,
            "deviations_by_node": deviations_by_node,
            "completion_percentage": 0,  # Can be calculated based on filled fields
            "query_time_ms": round(execution_time * 1000, 2)  # For debugging, can be removed in production
        }
    }


# Excel export endpoint
from fastapi.responses import StreamingResponse
import io

@router.get("/{study_id}/export/excel")
def export_study_to_excel(
    study_id: UUID,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export study to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        study = db.query(HazopStudy).filter(HazopStudy.id == study_id).first()
        if not study:
            raise HTTPException(status_code=404, detail="Study not found")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "HAZOP Analysis"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        
        # Headers
        headers = ["Node", "Parameter", "Guide Word", "Deviation", "Cause", "Likelihood", 
                   "Consequence", "Severity", "Safeguard", "Recommendation", "Risk Level"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Get all nodes
        nodes = db.query(HazopNode).filter(HazopNode.study_id == study_id).all()
        
        row_num = 2
        for node in nodes:
            deviations = db.query(Deviation).filter(Deviation.node_id == node.id).all()
            
            for deviation in deviations:
                causes = db.query(Cause).filter(Cause.deviation_id == deviation.id).all()
                
                if not causes:
                    # Deviation without causes
                    ws.cell(row=row_num, column=1).value = f"{node.node_number} - {node.node_name}"
                    ws.cell(row=row_num, column=2).value = deviation.parameter
                    ws.cell(row=row_num, column=3).value = deviation.guide_word
                    ws.cell(row=row_num, column=4).value = deviation.deviation_description
                    row_num += 1
                
                for cause in causes:
                    consequences = db.query(Consequence).filter(Consequence.cause_id == cause.id).all()
                    
                    if not consequences:
                        # Cause without consequences
                        ws.cell(row=row_num, column=1).value = f"{node.node_number} - {node.node_name}"
                        ws.cell(row=row_num, column=2).value = deviation.parameter
                        ws.cell(row=row_num, column=3).value = deviation.guide_word
                        ws.cell(row=row_num, column=4).value = deviation.deviation_description
                        ws.cell(row=row_num, column=5).value = cause.cause_description
                        ws.cell(row=row_num, column=6).value = cause.likelihood
                        row_num += 1
                    
                    for consequence in consequences:
                        safeguards = db.query(Safeguard).filter(Safeguard.consequence_id == consequence.id).all()
                        recommendations = db.query(Recommendation).filter(Recommendation.consequence_id == consequence.id).all()
                        impact = db.query(ImpactAssessment).filter(ImpactAssessment.consequence_id == consequence.id).first()
                        
                        # Write main row
                        ws.cell(row=row_num, column=1).value = f"{node.node_number} - {node.node_name}"
                        ws.cell(row=row_num, column=2).value = deviation.parameter
                        ws.cell(row=row_num, column=3).value = deviation.guide_word
                        ws.cell(row=row_num, column=4).value = deviation.deviation_description
                        ws.cell(row=row_num, column=5).value = cause.cause_description
                        ws.cell(row=row_num, column=6).value = cause.likelihood
                        ws.cell(row=row_num, column=7).value = consequence.consequence_description
                        ws.cell(row=row_num, column=8).value = consequence.severity
                        
                        # Safeguards (combine multiple)
                        if safeguards:
                            safeguard_text = "; ".join([s.safeguard_description for s in safeguards])
                            ws.cell(row=row_num, column=9).value = safeguard_text
                        
                        # Recommendations (combine multiple)
                        if recommendations:
                            rec_text = "; ".join([r.recommendation_description for r in recommendations])
                            ws.cell(row=row_num, column=10).value = rec_text
                        
                        # Risk level
                        if impact:
                            ws.cell(row=row_num, column=11).value = impact.risk_level
                            
                            # Color code by risk level
                            risk_colors = {
                                "Critical": "EF4444",
                                "High": "F59E0B",
                                "Medium": "FBBF24",
                                "Low": "10B981"
                            }
                            if impact.risk_level in risk_colors:
                                risk_cell = ws.cell(row=row_num, column=11)
                                risk_cell.fill = PatternFill(start_color=risk_colors[impact.risk_level], 
                                                            end_color=risk_colors[impact.risk_level], 
                                                            fill_type="solid")
                        
                        row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"{study.title.replace(' ', '_')}_HAZOP_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"Export error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
